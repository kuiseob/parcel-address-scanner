from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
from pathlib import Path
from typing import Optional


class AddressBook:
    """Excel 형식의 주소록 관리"""

    def __init__(self, filename: str = "주소록.xlsx"):
        self.filename = filename
        self.workbook = None
        self.sheet = None
        self._ensure_workbook()

    def _ensure_workbook(self):
        """워크북 생성 또는 로드"""
        path = Path(self.filename)

        if path.exists():
            self.workbook = load_workbook(path)
            self.sheet = self.workbook.active
        else:
            self._create_workbook()

    def _create_workbook(self):
        """새 워크북 생성"""
        self.workbook = Workbook()
        self.sheet = self.workbook.active
        self.sheet.title = "주소록"

        # 헤더 설정
        headers = ["번호", "주소", "등록일시", "메모"]
        self.sheet.append(headers)

        # 헤더 스타일
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for cell in self.sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border

        # 열 너비 설정
        self.sheet.column_dimensions['A'].width = 8
        self.sheet.column_dimensions['B'].width = 50
        self.sheet.column_dimensions['C'].width = 20
        self.sheet.column_dimensions['D'].width = 30

    def add_address(self, address: str, memo: str = "") -> bool:
        """
        주소 추가 (중복 방지)
        반환: 추가 성공 여부
        """
        if self._address_exists(address):
            return False

        row_num = self.sheet.max_row + 1
        self.sheet[f'A{row_num}'] = row_num - 1

        self.sheet[f'B{row_num}'] = address
        self.sheet[f'C{row_num}'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        self.sheet[f'D{row_num}'] = memo

        # 데이터 스타일
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for col in ['A', 'B', 'C', 'D']:
            cell = self.sheet[f'{col}{row_num}']
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=True)

        self.save()
        return True

    def _address_exists(self, address: str) -> bool:
        """주소 중복 확인"""
        for row in self.sheet.iter_rows(min_row=2, max_row=self.sheet.max_row, values_only=True):
            if row[1] and row[1].strip() == address.strip():
                return True
        return False

    def get_all_addresses(self) -> list[str]:
        """모든 주소 반환"""
        addresses = []
        for row in self.sheet.iter_rows(min_row=2, max_row=self.sheet.max_row, values_only=True):
            if row[1]:
                addresses.append(row[1])
        return addresses

    def save(self):
        """워크북 저장"""
        self.workbook.save(self.filename)

    def close(self):
        """워크북 종료"""
        if self.workbook:
            self.workbook.close()
